"""Tests for DompowerClient."""

import io
import json
from datetime import UTC, date, datetime
from typing import Any
from unittest.mock import MagicMock

import openpyxl
import pytest
from aiohttp import ClientSession
from aresponses import ResponsesMockServer

from dompower import (
    ApiError,
    DompowerClient,
    InvalidAuthError,
    RateLimitError,
)
from dompower.client import DOMINION_TIMEZONE

# ---------------------------------------------------------------------------
# Helpers to build in-memory Excel workbooks matching the Dominion API format
# ---------------------------------------------------------------------------


def _time_header(hour: int, minute: int) -> str:
    """Build a column header like '12:00 AM kWH'."""
    suffix = "AM" if hour < 12 else "PM"
    display_hour = hour % 12 or 12
    return f"{display_hour}:{minute:02d} {suffix} kWH"


# All 48 half-hour (hour, minute) pairs starting at midnight
_ALL_SLOTS = [(h, m) for h in range(24) for m in (0, 30)]
_ALL_HEADERS = ["Account No", "Recorder ID", "Date"] + [
    _time_header(h, m) for h, m in _ALL_SLOTS
]


def _make_excel(
    rows: list[tuple[date, list[float | None]]],
    *,
    generation_rows: list[tuple[date, list[float | None]]] | None = None,
) -> bytes:
    """Create an in-memory Excel workbook matching Dominion's format.

    Args:
        rows: List of (date, values) for consumption sheet.
            values is a list of up to 48 floats (one per half-hour slot).
        generation_rows: Optional list of (date, values) for generation sheet.

    Returns:
        Raw Excel bytes.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "kWH Consumption"

    # Header row
    ws.append(_ALL_HEADERS)

    for row_date, values in rows:
        row: list[Any] = ["ACCT", "METER", row_date]
        row.extend(values)
        # Pad with None to 48 columns if shorter
        while len(row) < 3 + 48:
            row.append(None)
        ws.append(row)

    if generation_rows:
        gs = wb.create_sheet("kWH Generation")
        gen_headers = [h.replace("kWH", "kW") for h in _ALL_HEADERS]
        gs.append(gen_headers)
        for row_date, values in generation_rows:
            row = ["ACCT", "METER", row_date]
            row.extend(values)
            while len(row) < 3 + 48:
                row.append(None)
            gs.append(row)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Client basics
# ---------------------------------------------------------------------------


class TestDompowerClient:
    """Core client tests (kept high-value only)."""

    async def test_token_callback(self) -> None:
        """Token callback is invoked with correct values on async_set_tokens."""
        received: dict[str, str] = {}

        def cb(access: str, refresh: str) -> None:
            received["access"] = access
            received["refresh"] = refresh

        async with ClientSession() as session:
            client = DompowerClient(session, token_update_callback=cb)
            await client.async_set_tokens("a", "r")

            assert received == {"access": "a", "refresh": "r"}

    async def test_set_tokens_values_retrievable(self) -> None:
        """After set_tokens, access and refresh tokens are retrievable."""
        async with ClientSession() as session:
            client = DompowerClient(session)
            assert not client.has_tokens

            await client.async_set_tokens("my_access", "my_refresh")
            assert client.has_tokens

            pair = client._token_manager.get_token_pair()
            assert pair is not None
            assert pair.access_token == "my_access"  # noqa: S105
            assert pair.refresh_token == "my_refresh"  # noqa: S105


# ---------------------------------------------------------------------------
# Token refresh
# ---------------------------------------------------------------------------


class TestTokenRefresh:
    """Tests for token refresh functionality."""

    async def test_refresh_tokens(
        self,
        aresponses: ResponsesMockServer,
        sample_tokens: dict[str, str],
        sample_refresh_response: dict[str, Any],
    ) -> None:
        """Successful refresh returns new tokens via callback."""
        aresponses.add(
            "prodsvc-dominioncip.smartcmobile.com",
            "/UsermanagementAPI/api/1/login/auth/refresh",
            "POST",
            response=aresponses.Response(
                body=json.dumps(sample_refresh_response),
                content_type="application/json",
            ),
        )

        callback_tokens: dict[str, str] = {}

        def token_callback(access: str, refresh: str) -> None:
            callback_tokens["access"] = access
            callback_tokens["refresh"] = refresh

        async with ClientSession() as session:
            client = DompowerClient(
                session,
                access_token=sample_tokens["access_token"],
                refresh_token=sample_tokens["refresh_token"],
                token_update_callback=token_callback,
            )

            await client.async_refresh_tokens()

            assert callback_tokens["access"] == "new_access_token_abc"
            assert callback_tokens["refresh"] == "new_refresh_token_xyz"

    async def test_refresh_without_tokens(self) -> None:
        """Refresh without tokens raises InvalidAuthError."""
        async with ClientSession() as session:
            client = DompowerClient(session)
            with pytest.raises(InvalidAuthError):
                await client.async_refresh_tokens()


# ---------------------------------------------------------------------------
# _parse_excel_usage — normal day
# ---------------------------------------------------------------------------


class TestParseExcelUsageNormalDay:
    """Test Excel parsing on a normal (non-DST) day."""

    def test_normal_day_produces_48_intervals(self) -> None:
        """A normal day with 48 values produces 48 IntervalUsageData."""
        day = date(2026, 1, 15)
        values = [round(0.5 + i * 0.01, 2) for i in range(48)]
        excel_bytes = _make_excel([(day, values)])

        async_session = MagicMock(spec=ClientSession)
        client = DompowerClient(async_session)
        result = client._parse_excel_usage(excel_bytes, day, day)

        assert len(result) == 48
        # First interval is midnight
        assert result[0].timestamp.hour == 0
        assert result[0].timestamp.minute == 0
        # Last interval is 23:30
        assert result[-1].timestamp.hour == 23
        assert result[-1].timestamp.minute == 30
        # Values preserved
        assert result[0].consumption == values[0]
        assert result[47].consumption == values[47]
        # All timestamps have Eastern timezone
        for r in result:
            assert r.timestamp.tzinfo == DOMINION_TIMEZONE

    def test_date_range_filtering(self) -> None:
        """Only intervals within the requested date range are returned."""
        day1 = date(2026, 1, 14)
        day2 = date(2026, 1, 15)
        day3 = date(2026, 1, 16)
        vals = [1.0] * 48
        excel_bytes = _make_excel(
            [
                (day1, vals),
                (day2, vals),
                (day3, vals),
            ]
        )

        client = DompowerClient(MagicMock(spec=ClientSession))
        result = client._parse_excel_usage(excel_bytes, day2, day2)

        assert len(result) == 48
        assert all(r.timestamp.date() == day2 for r in result)

    def test_none_values_skipped(self) -> None:
        """None values in the Excel are skipped, not zero-filled."""
        day = date(2026, 1, 15)
        values: list[float | None] = [1.0, None, 2.0] + [None] * 45
        excel_bytes = _make_excel([(day, values)])

        client = DompowerClient(MagicMock(spec=ClientSession))
        result = client._parse_excel_usage(excel_bytes, day, day)

        assert len(result) == 2
        assert result[0].consumption == 1.0
        assert result[1].consumption == 2.0

    def test_empty_workbook_returns_empty(self) -> None:
        """An Excel with no data rows returns an empty list."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(_ALL_HEADERS)  # header only
        buf = io.BytesIO()
        wb.save(buf)

        client = DompowerClient(MagicMock(spec=ClientSession))
        result = client._parse_excel_usage(
            buf.getvalue(), date(2026, 1, 1), date(2026, 12, 31)
        )
        assert result == []


# ---------------------------------------------------------------------------
# _parse_excel_usage — DST spring-forward
# ---------------------------------------------------------------------------


class TestParseExcelUsageDST:
    """Test Excel parsing around DST spring-forward (March 8, 2026)."""

    def _make_dst_day_excel(self) -> bytes:
        """Build Excel for March 8 2026 (spring-forward at 2 AM)."""
        day = date(2026, 3, 8)
        values = [round(0.5 + i * 0.01, 2) for i in range(48)]
        return _make_excel([(day, values)])

    def test_spring_forward_yields_46_intervals(self) -> None:
        """Spring-forward day has 46 intervals (gap times folded)."""
        excel_bytes = self._make_dst_day_excel()
        client = DompowerClient(MagicMock(spec=ClientSession))
        result = client._parse_excel_usage(
            excel_bytes, date(2026, 3, 8), date(2026, 3, 8)
        )

        # 2:00 and 2:30 AM are gap times → folded into 3:00/3:30 EDT
        assert len(result) == 46

    def test_spring_forward_no_duplicate_utc(self) -> None:
        """No duplicate UTC timestamps on spring-forward day."""
        excel_bytes = self._make_dst_day_excel()
        client = DompowerClient(MagicMock(spec=ClientSession))
        result = client._parse_excel_usage(
            excel_bytes, date(2026, 3, 8), date(2026, 3, 8)
        )

        utc_times = [r.timestamp.astimezone(UTC) for r in result]
        assert len(utc_times) == len(set(utc_times))

    def test_spring_forward_gap_values_folded(self) -> None:
        """Gap time consumption is added to the post-transition interval."""
        day = date(2026, 3, 8)
        # Set specific values so we can verify folding
        values = [0.0] * 48
        # Index 4 = 2:00 AM (gap), Index 5 = 2:30 AM (gap)
        # Index 6 = 3:00 AM, Index 7 = 3:30 AM
        values[4] = 1.0  # 2:00 AM — gap, folds into 3:00 AM EDT
        values[5] = 2.0  # 2:30 AM — gap, folds into 3:30 AM EDT
        values[6] = 3.0  # 3:00 AM
        values[7] = 4.0  # 3:30 AM

        excel_bytes = _make_excel([(day, values)])
        client = DompowerClient(MagicMock(spec=ClientSession))
        result = client._parse_excel_usage(excel_bytes, day, day)

        # Find the 3:00 AM EDT and 3:30 AM EDT intervals
        by_hour = {(r.timestamp.hour, r.timestamp.minute): r for r in result}
        # After spring-forward, 2:00 EST → 3:00 EDT in UTC round-trip
        # The gap consumption should be added to the 3:00 and 3:30 slots
        assert by_hour[(3, 0)].consumption == 4.0  # 1.0 + 3.0
        assert by_hour[(3, 30)].consumption == 6.0  # 2.0 + 4.0

    def test_normal_day_still_48_intervals(self) -> None:
        """A day that's NOT DST transition still produces 48 intervals."""
        day = date(2026, 3, 7)  # Day before spring-forward
        values = [1.0] * 48
        excel_bytes = _make_excel([(day, values)])

        client = DompowerClient(MagicMock(spec=ClientSession))
        result = client._parse_excel_usage(excel_bytes, day, day)
        assert len(result) == 48


# ---------------------------------------------------------------------------
# _parse_excel_usage — generation data
# ---------------------------------------------------------------------------


class TestParseExcelUsageGeneration:
    """Test consumption + generation recombination."""

    def test_generation_merged_into_intervals(self) -> None:
        """Generation data from second sheet is merged with consumption."""
        day = date(2026, 6, 15)
        consumption = [1.0] * 48
        generation = [0.0] * 48
        generation[24] = 0.5  # noon: 0.5 kWh generated
        generation[25] = 0.3  # 12:30 PM

        excel_bytes = _make_excel(
            [(day, consumption)],
            generation_rows=[(day, generation)],
        )

        client = DompowerClient(MagicMock(spec=ClientSession))
        result = client._parse_excel_usage(excel_bytes, day, day)

        assert len(result) == 48
        noon = next(
            r for r in result if r.timestamp.hour == 12 and r.timestamp.minute == 0
        )
        assert noon.generation == 0.5
        assert noon.consumption == 1.0

        half = next(
            r for r in result if r.timestamp.hour == 12 and r.timestamp.minute == 30
        )
        assert half.generation == 0.3

    def test_no_generation_sheet_defaults_zero(self) -> None:
        """Without a generation sheet, generation defaults to 0."""
        day = date(2026, 6, 15)
        excel_bytes = _make_excel([(day, [1.0] * 48)])

        client = DompowerClient(MagicMock(spec=ClientSession))
        result = client._parse_excel_usage(excel_bytes, day, day)

        assert all(r.generation == 0.0 for r in result)


# ---------------------------------------------------------------------------
# _parse_bill_forecast / _parse_date
# ---------------------------------------------------------------------------


class TestParseBillForecast:
    """Test bill forecast parsing and derived_rate."""

    def _sample_response(self) -> dict[str, Any]:
        return {
            "status": {"type": "success", "code": 200},
            "data": {
                "lastBill": {
                    "charges": 150.00,
                    "usage": 1200.0,
                    "billPerdStDate": "01/01/2026 00:00:00",
                    "billPerdEdDate": "01/31/2026 00:00:00",
                },
                "billperdstdate": "02/01/2026 00:00:00",
                "billperdeddate": "02/28/2026 00:00:00",
                "currentUsageKwh": 450.5,
                "tou": "N",
            },
        }

    def test_parse_bill_forecast_full(self) -> None:
        """Parse a complete bill forecast response."""
        client = DompowerClient(MagicMock(spec=ClientSession))
        forecast = client._parse_bill_forecast(self._sample_response())

        assert forecast.last_bill.charges == 150.00
        assert forecast.last_bill.usage == 1200.0
        assert forecast.last_bill.period_start == date(2026, 1, 1)
        assert forecast.last_bill.period_end == date(2026, 1, 31)
        assert forecast.current_period_start == date(2026, 2, 1)
        assert forecast.current_period_end == date(2026, 2, 28)
        assert forecast.current_usage_kwh == 450.5
        assert forecast.is_tou is False

    def test_derived_rate(self) -> None:
        """derived_rate calculates $/kWh from last bill."""
        client = DompowerClient(MagicMock(spec=ClientSession))
        forecast = client._parse_bill_forecast(self._sample_response())

        assert forecast.derived_rate == pytest.approx(150.0 / 1200.0)

    def test_derived_rate_zero_usage(self) -> None:
        """derived_rate returns None when usage is zero."""
        resp = self._sample_response()
        resp["data"]["lastBill"]["usage"] = 0

        client = DompowerClient(MagicMock(spec=ClientSession))
        forecast = client._parse_bill_forecast(resp)
        assert forecast.derived_rate is None

    def test_missing_dates_default_to_today(self) -> None:
        """Missing current period dates default to today."""
        resp = self._sample_response()
        del resp["data"]["billperdstdate"]
        del resp["data"]["billperdeddate"]

        client = DompowerClient(MagicMock(spec=ClientSession))
        forecast = client._parse_bill_forecast(resp)

        assert forecast.current_period_start == date.today()
        assert forecast.current_period_end == date.today()

    def test_parse_date_valid(self) -> None:
        """_parse_date handles MM/DD/YYYY HH:MM:SS format."""
        client = DompowerClient(MagicMock(spec=ClientSession))
        assert client._parse_date("03/15/2026 00:00:00") == date(2026, 3, 15)

    def test_parse_date_none(self) -> None:
        """_parse_date returns None for None input."""
        client = DompowerClient(MagicMock(spec=ClientSession))
        assert client._parse_date(None) is None

    def test_parse_date_invalid(self) -> None:
        """_parse_date returns None for unparseable input."""
        client = DompowerClient(MagicMock(spec=ClientSession))
        assert client._parse_date("not-a-date") is None

    def test_tou_flag(self) -> None:
        """is_tou is True when tou='Y'."""
        resp = self._sample_response()
        resp["data"]["tou"] = "Y"

        client = DompowerClient(MagicMock(spec=ClientSession))
        forecast = client._parse_bill_forecast(resp)
        assert forecast.is_tou is True


# ---------------------------------------------------------------------------
# _async_request — 401 retry
# ---------------------------------------------------------------------------


class TestAsyncRequest401Retry:
    """Test the 401 → refresh → retry logic."""

    async def test_401_triggers_refresh_and_retry(
        self,
        aresponses: ResponsesMockServer,
        sample_refresh_response: dict[str, Any],
    ) -> None:
        """A 401 response triggers token refresh and a successful retry."""
        success_body = {"status": {"code": 200}, "data": {"result": "ok"}}

        # First request → 401
        aresponses.add(
            "prodsvc-dominioncip.smartcmobile.com",
            "/Service/api/1/test",
            "GET",
            response=aresponses.Response(status=401),
        )
        # Token refresh
        aresponses.add(
            "prodsvc-dominioncip.smartcmobile.com",
            "/UsermanagementAPI/api/1/login/auth/refresh",
            "POST",
            response=aresponses.Response(
                body=json.dumps(sample_refresh_response),
                content_type="application/json",
            ),
        )
        # Retry → 200
        aresponses.add(
            "prodsvc-dominioncip.smartcmobile.com",
            "/Service/api/1/test",
            "GET",
            response=aresponses.Response(
                body=json.dumps(success_body),
                content_type="application/json",
            ),
        )

        async with ClientSession() as session:
            client = DompowerClient(
                session,
                access_token="old_access",  # noqa: S106
                refresh_token="old_refresh",  # noqa: S106
            )
            result = await client._async_request("GET", "/Service/api/1/test")

        assert result["data"]["result"] == "ok"


# ---------------------------------------------------------------------------
# _handle_response error paths
# ---------------------------------------------------------------------------


class TestHandleResponseErrors:
    """Test error handling in _handle_response."""

    async def test_429_raises_rate_limit(
        self,
        aresponses: ResponsesMockServer,
    ) -> None:
        """429 response raises RateLimitError with retry_after."""
        aresponses.add(
            "prodsvc-dominioncip.smartcmobile.com",
            "/Service/api/1/test",
            "GET",
            response=aresponses.Response(
                status=429,
                headers={"Retry-After": "60"},
            ),
        )

        async with ClientSession() as session:
            client = DompowerClient(
                session,
                access_token="tok",  # noqa: S106
                refresh_token="ref",  # noqa: S106
            )
            # Bypass token expiry check
            client._token_manager._token_expires_at = datetime(2099, 1, 1, tzinfo=UTC)

            with pytest.raises(RateLimitError) as exc_info:
                await client._async_request("GET", "/Service/api/1/test")

            assert exc_info.value.retry_after == 60

    async def test_4xx_raises_api_error(
        self,
        aresponses: ResponsesMockServer,
    ) -> None:
        """4xx responses raise ApiError with status code."""
        aresponses.add(
            "prodsvc-dominioncip.smartcmobile.com",
            "/Service/api/1/test",
            "GET",
            response=aresponses.Response(status=403, body="Forbidden"),
        )

        async with ClientSession() as session:
            client = DompowerClient(
                session,
                access_token="tok",  # noqa: S106
                refresh_token="ref",  # noqa: S106
            )
            client._token_manager._token_expires_at = datetime(2099, 1, 1, tzinfo=UTC)

            with pytest.raises(ApiError) as exc_info:
                await client._async_request("GET", "/Service/api/1/test")

            assert exc_info.value.status_code == 403

    async def test_api_level_error_in_body(
        self,
        aresponses: ResponsesMockServer,
    ) -> None:
        """API-level error in JSON body raises ApiError."""
        body = {
            "status": {
                "type": "error",
                "code": 500,
                "message": "Internal failure",
                "error": True,
            },
        }
        aresponses.add(
            "prodsvc-dominioncip.smartcmobile.com",
            "/Service/api/1/test",
            "GET",
            response=aresponses.Response(
                body=json.dumps(body),
                content_type="application/json",
            ),
        )

        async with ClientSession() as session:
            client = DompowerClient(
                session,
                access_token="tok",  # noqa: S106
                refresh_token="ref",  # noqa: S106
            )
            client._token_manager._token_expires_at = datetime(2099, 1, 1, tzinfo=UTC)

            with pytest.raises(ApiError, match="Internal failure"):
                await client._async_request("GET", "/Service/api/1/test")


# ---------------------------------------------------------------------------
# _parse_accounts_response
# ---------------------------------------------------------------------------


class TestParseAccountsResponse:
    """Test account parsing from API JSON."""

    def _sample_accounts_response(self) -> dict[str, Any]:
        return {
            "data": [
                {
                    "firstName": "Jane",
                    "lastName": "Doe",
                    "primaryEmail": "jane@example.com",
                    "zbpMaintRegEnroll_nav": {
                        "results": [
                            {
                                "account": "111222333444",
                                "premiseNumber": "P001",
                                "accountActive": "Y",
                                "default": True,
                                "nickName": "Home",
                                "ebillStatus": "Y",
                                "serviceAddress": {
                                    "street": "Main St",
                                    "houseNum1": "123",
                                    "city": "Richmond",
                                    "state": "VA",
                                    "zipCode": "23219",
                                    "country": "US",
                                },
                                "conDev": [
                                    {
                                        "device": "DEV001",
                                        "contract": "CON001",
                                        "isActive": "Y",
                                        "amiFlag": True,
                                        "netMetering": "Y",
                                    }
                                ],
                            },
                            {
                                "account": "555666777888",
                                "premiseNumber": "P002",
                                "accountActive": "N",
                                "default": False,
                                "nickName": "",
                                "ebillStatus": "N",
                                "closingDate": "12/01/2025 00:00:00",
                                "serviceAddress": {
                                    "street": "Oak Ave",
                                    "houseNum1": "456",
                                    "city": "Norfolk",
                                    "state": "VA",
                                    "zipCode": "23510",
                                },
                                "conDev": [
                                    {
                                        "device": "DEV002",
                                        "contract": "CON002",
                                        "isActive": "N",
                                        "amiFlag": False,
                                    }
                                ],
                            },
                        ]
                    },
                }
            ]
        }

    def test_active_only_by_default(self) -> None:
        """Only active accounts returned when include_inactive=False."""
        client = DompowerClient(MagicMock(spec=ClientSession))
        accounts = client._parse_accounts_response(
            self._sample_accounts_response(), include_inactive=False
        )

        assert len(accounts) == 1
        assert accounts[0].account_number == "111222333444"
        assert accounts[0].is_active is True

    def test_include_inactive(self) -> None:
        """All accounts returned when include_inactive=True."""
        client = DompowerClient(MagicMock(spec=ClientSession))
        accounts = client._parse_accounts_response(
            self._sample_accounts_response(), include_inactive=True
        )

        assert len(accounts) == 2

    def test_account_fields_parsed(self) -> None:
        """Account fields (address, meters, nickname, etc.) parsed correctly."""
        client = DompowerClient(MagicMock(spec=ClientSession))
        accounts = client._parse_accounts_response(
            self._sample_accounts_response(), include_inactive=True
        )

        acct = accounts[0]
        assert acct.nickname == "Home"
        assert acct.is_default is True
        assert acct.ebill_enrolled is True
        assert str(acct.service_address) == "123 Main St, Richmond, VA 23219"

        meter = acct.meters[0]
        assert meter.device_id == "DEV001"
        assert meter.has_ami is True
        assert meter.net_metering is True

    def test_inactive_account_fields(self) -> None:
        """Inactive account has closing_date and net_metering=None."""
        client = DompowerClient(MagicMock(spec=ClientSession))
        accounts = client._parse_accounts_response(
            self._sample_accounts_response(), include_inactive=True
        )

        inactive = accounts[1]
        assert inactive.is_active is False
        assert inactive.closing_date == date(2025, 12, 1)
        assert inactive.nickname is None  # empty string → None
        assert inactive.meters[0].net_metering is None

    def test_empty_response(self) -> None:
        """Empty data array returns empty list."""
        client = DompowerClient(MagicMock(spec=ClientSession))
        result = client._parse_accounts_response({"data": []}, include_inactive=True)
        assert result == []
